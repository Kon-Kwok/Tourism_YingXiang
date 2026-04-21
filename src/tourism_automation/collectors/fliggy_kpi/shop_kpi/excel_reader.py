"""Excel文件读取器 - 解析导出的KPI数据"""

import openpyxl
from pathlib import Path
from typing import Dict, List, Any
from datetime import datetime


class ShopKpiExcelReader:
    """店铺KPI Excel文件读取器"""

    @staticmethod
    def read_excel(file_path: str) -> Dict[str, Any]:
        """读取Excel文件

        Args:
            file_path: Excel文件路径

        Returns:
            Dict: 解析后的数据
        """
        wb = openpyxl.load_workbook(file_path, data_only=True)

        result = {
            'file_path': file_path,
            'file_name': Path(file_path).name,
            'sheets': {},
            'sheet_names': wb.sheetnames
        }

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            sheet_data = {
                'name': sheet_name,
                'dimensions': f"{ws.max_row} rows x {ws.max_column} cols",
                'headers': [],
                'rows': []
            }

            # 读取所有行
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                # 跳过完全空行
                if not any(cell is not None for cell in row):
                    continue

                row_data = [cell if cell is not None else '' for cell in row]

                # 第一行作为表头
                if row_idx == 1:
                    sheet_data['headers'] = row_data
                else:
                    sheet_data['rows'].append(row_data)

            result['sheets'][sheet_name] = sheet_data

        return result

    @staticmethod
    def print_summary(data: Dict[str, Any]):
        """打印数据摘要

        Args:
            data: read_excel返回的数据
        """
        print("=" * 70)
        print(f"文件: {data['file_name']}")
        print(f"路径: {data['file_path']}")
        print(f"工作表数量: {len(data['sheet_names'])}")
        print("=" * 70)

        for sheet_name, sheet_data in data['sheets'].items():
            print(f"\n【工作表】: {sheet_name}")
            print(f"维度: {sheet_data['dimensions']}")
            print(f"列数: {len(sheet_data['headers'])}")
            print(f"数据行数: {len(sheet_data['rows'])}")

            print(f"\n表头:")
            for i, header in enumerate(sheet_data['headers']):
                print(f"  {i+1}. {header}")

            if len(sheet_data['rows']) > 0:
                print(f"\n前5行数据:")
                for i, row in enumerate(sheet_data['rows'][:5], 1):
                    print(f"  Row {i}: {' | '.join(str(v) for v in row)}")

                if len(sheet_data['rows']) > 5:
                    print(f"  ... 还有 {len(sheet_data['rows']) - 5} 行")

    @staticmethod
    def to_dict(data: Dict[str, Any]) -> List[Dict[str, Any]]:
        """将Excel数据转换为字典列表

        Args:
            data: read_excel返回的数据

        Returns:
            List[Dict]: 字典列表
        """
        result = []

        for sheet_name, sheet_data in data['sheets'].items():
            headers = sheet_data['headers']

            for row in sheet_data['rows']:
                row_dict = {
                    '_sheet': sheet_name,
                    '_row_number': len(result) + 1
                }

                for i, header in enumerate(headers):
                    if i < len(row):
                        row_dict[header] = row[i]
                    else:
                        row_dict[header] = None

                result.append(row_dict)

        return result


if __name__ == "__main__":
    # 测试读取
    import sys

    if len(sys.argv) > 1:
        file_path = sys.argv[1]
    else:
        # 默认读取已知的员工KPI文件
        file_path = "/home/kk/下载/自定义报表_人均日接入_下单优先判定_2026-04-19至2026-04-19 (1).xlsx"

    reader = ShopKpiExcelReader()
    data = reader.read_excel(file_path)
    reader.print_summary(data)

    print("\n" + "=" * 70)
    print("转换为字典格式:")
    print("=" * 70)

    dict_data = reader.to_dict(data)
    for i, row in enumerate(dict_data[:3], 1):
        print(f"\n记录 {i}:")
        for key, value in row.items():
            if not key.startswith('_'):
                print(f"  {key}: {value}")
